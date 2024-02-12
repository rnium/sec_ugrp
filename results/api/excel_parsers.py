import openpyxl
from io import BytesIO
from results.utils import get_letter_grade
from datetime import datetime

def parse_student_data(wb: openpyxl.Workbook):
    student_data = {}
    rows = list((wb['STUDENT_INFO']).rows)
    headers = [cell.value for cell in rows[0]]
    key_idx = headers.index('FIELDS')
    value_idx = headers.index('VALUES')
    for row in rows[1:]:
        value = row[value_idx].value
        if type(value) == datetime:
            value = value.strftime("%Y")
        key = row[key_idx].value
        key = key.strip() if type(value) == str else key
        student_data[key] = value.strip() if type(value) == str else value
    return student_data

def parse_semester_data(wb: openpyxl.Workbook, sheetname):
    rows = list((wb[sheetname]).rows)
    headers = [cell.value.strip() if cell.value else None for cell in rows[0]]
    course_code_idx = headers.index('course_code')
    course_title_idx = headers.index('course_title')
    course_credit_idx = headers.index('course_credit')
    grade_point_idx = headers.index('grade_point')
    sem_data = {'courses': []}
    
    for row in rows[1:]:
        if row[course_code_idx].value is None:
            break
        sem_data['courses'].append({
            'code': row[course_code_idx].value,
            'title': row[course_title_idx].value,
            'credit': row[course_credit_idx].value,
            'gp': row[grade_point_idx].value,
            'lg': get_letter_grade(row[grade_point_idx].value),
        })
    info_idx = headers.index('SEMESTER_INFO')
    for row in rows[1:8]:
        key = row[info_idx].value
        value = row[info_idx+1].value
        key = key.strip() if type(key) == str else key
        if type(value) == datetime:
            value = value.strftime("%B %Y")
        sem_data[key] = value
    return sem_data
    


def parse_customdoc_excel(excel_file):
    buffer = BytesIO(excel_file.read())
    wb = openpyxl.load_workbook(buffer)
    parsed_data = {
        'student_data': parse_student_data(wb),
        'years': {
            1: {},
            2: {},
            3: {},
            4: {},
        }
    }
    get_year = lambda n: n%2 + n//2
    get_year_semester = lambda n: (1 if n%2 else 2)
    for sheetname in wb.sheetnames:
        if sheetname.startswith("SEM"):
            semester_data = parse_semester_data(wb, sheetname)
            if len(semester_data['courses']) >= 1:
                sem_num = int(sheetname.split("_")[1])
                parsed_data['years'][get_year(sem_num)][get_year_semester(sem_num)] = semester_data
    return parsed_data


def parse_course_sustdocs_excel(excel_file):
    buffer = BytesIO(excel_file.read())
    wb = openpyxl.load_workbook(buffer)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.rows)
    header = [cell.value.lower().strip() if cell.value is not None else None for cell in rows[0]]
    data_rows = rows[1:]
    parsed_data = {}
    fields_col_idx = header.index('field_name')
    value_col_idx = header.index('value')
    registrations_col_idx = header.index('additional_registrations')
    total_score_col_idx = header.index('total_score')
    expelled_registrations_col_idx = header.index('expelled_registrations')
    for i in range(8):
        parsed_data[data_rows[i][fields_col_idx].value] = data_rows[i][value_col_idx].value 
    parsed_data['additional_entries'] = []
    parsed_data['expelled_registrations'] = []
    for row in data_rows:
        row_data = [row[registrations_col_idx].value, row[total_score_col_idx].value]
        if any(row_data):
            parsed_data['additional_entries'].append(row_data)
        if expelled:=row[expelled_registrations_col_idx].value:
            parsed_data['expelled_registrations'].append(expelled)
    return parsed_data
 