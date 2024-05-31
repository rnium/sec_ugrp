from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from account.models import StudentAccount
from results.models import (Session, Course,
                            CourseResult, Semester, SemesterEnroll,
                            Department, StudentPoint, StudentCustomDocument, StudentAcademicData)
from results.utils import get_letter_grade
from django.forms.models import model_to_dict
from django.core.exceptions import ValidationError
from django.http import Http404
from io import StringIO, BytesIO
from . import excel_parsers
from results.pdf_generators import customdoc_generator
from results.tasks import update_student_accounts
from django.template.loader import render_to_string
from django.urls import reverse
from django.templatetags.static import static
import openpyxl


def create_course_enrollments(semester: Semester):
    object_prototypes = []
    for student in semester.session.studentaccount_set.all():
        object_prototypes.append(SemesterEnroll(student=student, semester=semester))
    
    SemesterEnroll.objects.bulk_create(object_prototypes)
    
def add_course_to_enrollments(course: Course):
    enrollments = SemesterEnroll.objects.filter(semester=course.semester)
    for enroll in enrollments:
        enroll.courses.add(course)

def create_course_results(course: Course):
    object_prototypes = []
    for enroll in course.semester.semesterenroll_set.all():
        object_prototypes.append(CourseResult(student=enroll.student, course=course))
    
    CourseResult.objects.bulk_create(object_prototypes)
    

def add_course_and_create_courseresults(enroll: SemesterEnroll):
    courses = enroll.semester.course_set.all()
    object_prototypes = []
    for course in courses:
        enroll.courses.add(course)
        object_prototypes.append(CourseResult(student=enroll.student, course=course))
    CourseResult.objects.bulk_create(object_prototypes)


def format_render_config(request):
    data = {
        'render_config': request.data.get('render_config', {'tabulation_title':'', 'tabulation_exam_time':''}),
        'footer_data': {
            'chairman': request.data['footer_data_raw']['chairman'],
            'controller': request.data['footer_data_raw']['controller'],
            # member and tabulators
        }
    }
    members = request.data['footer_data_raw']['committee']
    for i, member in enumerate(members):
        data['footer_data'][f'committee_mem_{i+1}'] = member
    tabulators = request.data['footer_data_raw']['tabulators']
    for i, tabulator in enumerate(tabulators):
        data['footer_data'][f'tabulator_{i+1}'] = tabulator
    return data


def is_confirmed_user(request, username):
    password = request.data.get('password')
    user = authenticate(request, username=username, password=password)
    return ((user is not None) and (request.user == user))

def course_sorting_key(course: Course, dept_name):
    dept_priority = 0
    coursecode_first_part = course.code.split(' ')[0].lower().strip()
    if coursecode_first_part == dept_name.lower():
        dept_priority = 1
    return (course.course_credit, dept_priority)

def sort_courses(courses, dept_name):
    sorted_courses = sorted(courses, key=lambda course: course_sorting_key(course, dept_name), reverse=True)
    return sorted_courses

def create_backup(dept: Department, session_id):
    backup_data = {
        "dept": dept.name,
        "single_batch_type": session_id is not None,
        "sessions": []
    }
    if session_id:
        sessions = Session.objects.filter(id=session_id, dept=dept)
    else:
        sessions = Session.objects.filter(dept=dept)
    for session in sessions:
        session_data = {
            'session_meta' : model_to_dict(session),
            'students': [],
            'semesters': [],
            'previous_point': None
        }
        for student in StudentAccount.objects.filter(session=session):
            student_data = model_to_dict(student)
            student_data.pop('profile_picture')
            session_data['students'].append(student_data)
        if hasattr(session, 'previouspoint'):
            session_data['previous_point'] = {
                'prevpoint_meta': model_to_dict(session.previouspoint),
                'student_points': []
            }
            for spoint in session.previouspoint.studentpoint_set.all():
                session_data['previous_point']['student_points'].append(model_to_dict(spoint))
        semesters = Semester.objects.filter(session=session)
        for semester in semesters:
            # items: document, courses, enrollments
            semester_doc_dict = None
            if hasattr(semester, 'semesterdocument'):
                semester_doc_dict = model_to_dict(semester.semesterdocument)
                semester_doc_dict.pop('tabulation_sheet')
                semester_doc_dict.pop('tabulation_thumbnail')
                semester_doc_dict['tabulation_sheet_render_time'] = None
                semester_doc_dict['tabulation_sheet_render_by'] = None
                print(semester_doc_dict, flush=1)
            semester_data = {
                'semester_meta': model_to_dict(semester),
                'semester_doc': semester_doc_dict,
                'courses': [],
                'enrolls': []
            }
            semester_drop_courses = semester_data['semester_meta']['drop_courses']
            semester_drop_courses_ids = [course.id for course in semester_drop_courses]
            semester_data['semester_meta']['drop_courses'] = semester_drop_courses_ids
            # adding drop course results in the semester if creating backup for session
            if session_id:
                semester_drop_courses_courseresults_data = {} 
                for d_course in semester_drop_courses:
                    drop_course_results = d_course.courseresult_set.filter(student__session=semester.session)
                    drop_course_courseresults_data = []
                    for result in drop_course_results:
                        drop_course_courseresults_data.append(model_to_dict(result))
                    semester_drop_courses_courseresults_data[d_course.id] = drop_course_courseresults_data
                semester_data['drop_course_courseresults'] = semester_drop_courses_courseresults_data
            if (semester.added_by):
                semester_data['semester_meta']['added_by'] = semester.added_by.username
            courses = Course.objects.filter(semester=semester).order_by('id')
            for course in courses:
                metadata = model_to_dict(course)
                metadata['added_in'] = metadata['added_in'].isoformat()
                course_data = {
                    'course_meta': metadata,
                    'course_results': []
                }
                if course.added_by:
                    course_data['course_meta']['added_by'] = course.added_by.username
                course_results = CourseResult.objects.filter(course=course)
                for result in course_results:
                    course_data['course_results'].append(model_to_dict(result))
                semester_data['courses'].append(course_data)
            enrolls = SemesterEnroll.objects.filter(semester=semester)
            for enroll in enrolls:
                enroll_data = model_to_dict(enroll)
                enrolled_courses = [course.id for course in enroll_data['courses']]
                enroll_data['courses'] = enrolled_courses
                semester_data['enrolls'].append(enroll_data)
                
            session_data['semesters'].append(semester_data)
        backup_data['sessions'].append(session_data)
    
    return backup_data
    
def delete_all_of_dept(dept: Department):
    sessions = dept.session_set.all()
    sessions.delete()

def get_obj_count(sessions_data):
    obj_count = len(sessions_data)
    for session in sessions_data:
        obj_count += len(session['students'])        
        obj_count += len(session['semesters'])
        if session.get('previous_point'):
            obj_count += len(session['previous_point']['student_points'])
        for semester in session['semesters']:
            obj_count += len(semester['courses'])
            obj_count += len(semester['enrolls'])
            obj_count += len(semester['semester_meta']['drop_courses'])
            obj_count += sum(len(course['course_results']) for course in semester['courses'])
            obj_count += sum(len(enroll['courses']) for enroll in semester['enrolls'])
            drop_courses_courseresults = semester.get('drop_course_courseresults')
            if drop_courses_courseresults:
                for drop_course in drop_courses_courseresults:
                    obj_count += len(drop_courses_courseresults[drop_course])
            
    return obj_count

def check_session_dependancy(session_data):
    drop_courses_ids = []
    for semester_data in session_data['semesters']:
        drop_courses_ids.extend(semester_data['semester_meta']['drop_courses'])
    for course_id in drop_courses_ids:
        try:
            Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return (False, f"Course with id: <{course_id}> does not exists")
    return (True, "All Dependancy Exists")
                
            
def parse_gradesheet_excel(excel_file, form_data, num_semesters):    
    buffer = BytesIO(excel_file.read())
    wb = openpyxl.load_workbook(buffer)
    parsed_data = {}
    if (num_semesters > 2):
        raise ValidationError("Number of semesters exceeds limit")
    for i in range(num_semesters):
        sheet = wb[wb.sheetnames[i]]
        rows = list(sheet.rows)
        header = [cell.value.lower().strip() if cell.value is not None else None for cell in rows[0]]
        data_rows = rows[1:]
        semester_key = f'semester_{i+1}'
        parsed_data[semester_key] = {
            'courses': []
        }
        code_idx = header.index('course_code')
        title_idx = header.index('course_title')
        credit_idx = header.index('course_credit')
        gp_idx = header.index('grade_point')
        total_credits = 0
        for row in range(len(data_rows)):
            course_credit = float(data_rows[row][credit_idx].value)
            course_data = {
                'code': data_rows[row][code_idx].value,
                'title': data_rows[row][title_idx].value,
                'credit': course_credit,
                'gp': float(data_rows[row][gp_idx].value),
                'lg': get_letter_grade(float(data_rows[row][gp_idx].value)),
            }
            parsed_data[semester_key]['courses'].append(course_data)
            total_credits += course_credit
        parsed_data[semester_key]['semester_credits'] = total_credits
        parsed_data[semester_key]['semester_gp'] = float(data_rows[0][header.index('semester_gp')].value)
        parsed_data[semester_key]['cumulative_credits'] = data_rows[0][header.index('cumulative_credits')].value
        parsed_data[semester_key]['cumulative_gp'] = data_rows[0][header.index('cumulative_gp')].value
        parsed_data[semester_key]['cumulative_lg'] = data_rows[0][header.index('cumulative_lg')].value
    parsed_data['semester_1']['year'] = form_data['first_sem_year']
    parsed_data['semester_1']['year_semester'] = form_data['first_sem_number']
    parsed_data['semester_1']['held_in'] = form_data['first_sem_held_in']
    if form_data['num_semesters'] == 2:
        parsed_data['semester_2']['year'] = form_data['second_sem_year']
        parsed_data['semester_2']['year_semester'] = form_data['second_sem_number']
        parsed_data['semester_2']['held_in'] = form_data['second_sem_held_in']
    return parsed_data


def parse_and_save_customdoc(excel_file, admin_name, user):
    data = excel_parsers.parse_customdoc_excel(excel_file)
    cdoc = customdoc_generator.process_and_save_customdoc_data(data, admin_name)
    return cdoc

def rank_students(students):
    return sorted(students, key=lambda student: (-student.credits_completed, -student.raw_cgpa))


def createStudentPointsFromExcel(excel_file, prevPoint, session):
    buffer = BytesIO(excel_file.read())
    wb = openpyxl.load_workbook(buffer)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.rows)
    header = [cell.value.lower().strip() if type(cell.value) == str else cell.value for cell in rows[0]]
    data_rows = rows[1:]
    logs = {
        'success': 0,
        'errors': {
            'parse_errors': [],
            'save_errors': []
        }
    }
    try:
        reg_col_idx = header.index('reg')
        credits_col_idx = header.index('credits')
        cgpa_col_idx = header.index('cgpa')
    except Exception as e:
        logs['errors']['parse_errors'].append(f'Error: {e}')
        data_rows = [] # Stopping the loop
    update_list = [] # For updating stats of studentAccounts
    for r in range(len(data_rows)):
        try:
            registration = int(data_rows[r][reg_col_idx].value)
            student_ac = StudentAccount.objects.get(registration=registration, session=session)
        except Exception as e:
            logs['errors']['parse_errors'].append(f'row: {r+2}. Errors: [Unmatched registration no.]')
            continue
        total_credits = data_rows[r][credits_col_idx].value
        if total_credits is None:
            logs['errors']['parse_errors'].append(f'registration: {registration}. Errors: [Total credits cannot be empty]')
            continue
        cgpa = data_rows[r][cgpa_col_idx].value
        if cgpa is None:
            logs['errors']['parse_errors'].append(f'registration: {registration}. Errors: [CGPA cannot be empty]')
            continue
        student_entrypoint_kwargs = {
            'prev_point': prevPoint,
            'student': student_ac,
            'total_credits': total_credits,
            'total_points': (total_credits * cgpa)
        }
        existing_point = StudentPoint.objects.filter(student__registration=registration).first()
        if existing_point:
            existing_point.total_credits = student_entrypoint_kwargs['total_credits']
            existing_point.total_points = student_entrypoint_kwargs['total_points']
            existing_point.save()
            logs['success'] += 1
            update_list.append(registration)
            continue
        try:
            StudentPoint.objects.create(**student_entrypoint_kwargs)
        except Exception as e:
            logs['errors']['save_errors'].append(f'Registration: {registration}. Errors: {e}')
            continue
        logs['success'] += 1
        update_list.append(registration)
    
    logs['has_parse_errors'] = bool(len(logs['errors']['parse_errors']))    
    logs['has_save_errors'] = bool(len(logs['errors']['save_errors']))
    update_student_accounts.delay(update_list)
    summary = render_to_string('results/components/excel_summary_list.html', context={'logs': logs})
    return summary
    # return JsonResponse({'status':'Complete', 'summary':summary})
    

def get_or_create_entry_for_carryCourse(student, course):
    # if semester is None:
    #     semester = course.semester
    enrollment = SemesterEnroll.objects.filter(student=student, semester__is_running=True).order_by('-semester__semester_no').first()
    print(enrollment.semester.semester_code, flush=1)
    if (enrollment is None) or (student is None):
        return None
    course_res, created = CourseResult.objects.get_or_create(student=student, course=course, is_drop_course=True)
    if created:
        print(f'Adding it, enrollid: {enrollment.id}', flush=1)
        enrollment.courses.add(course)
    return course_res

def try_get_carrycourse_retake_of(course_result):
    res = CourseResult.objects.filter(
        student=course_result.student, 
        course__code=course_result.course.code,
        letter_grade='F'
    ).exclude(id=course_result.id)
    return res.first()


def save_academic_studentdata(all_data):
    prototypes = []
    for data in all_data:
        session = data.get('session')
        registration = data['registration']
        academic_dat = StudentAcademicData.objects.filter(registration=registration).first()
        if academic_dat is not None:
            academic_dat.session_code = session
            academic_dat.data = data
            academic_dat.save()
        else:
            prototypes.append(StudentAcademicData(registration=registration, session_code=session, data=data))
    StudentAcademicData.objects.bulk_create(prototypes)
    

def academic_student_data(registration, dept_name):
    student_ac = StudentAccount.objects.filter(pk=registration, session__dept__name=dept_name).first()
    student_acadoc = StudentAcademicData.objects.filter(registration=registration).first()
    if not student_ac and not student_acadoc:
        raise Http404()
    response_data = {
        'registration': registration,
        'testimonial_url': reverse('results:download_testimonial', args=(registration,)),
        'coursemedium_url': reverse('results:download_coursemediumcert', args=(registration,)),
        'appearedCert_url': reverse('results:download_appeared_cert', args=(registration,)),
    }
    if student_acadoc:
        response_data['name'] = student_acadoc.data.get('name')
        response_data['dept'] = student_acadoc.data.get('dept')
        response_data['session'] = student_acadoc.session_code
        response_data['profile_picture_url'] = static('results/images/blank-dp.svg')
    else:
        student_data = {
            'name': student_ac.student_name,
            'dept': student_ac.session.dept.fullname,
            'session': student_ac.session.session_code,
            'profile_picture_url': student_ac.avatar_url
        }
        response_data = {**response_data, **student_data}
    return (student_ac, response_data)
        