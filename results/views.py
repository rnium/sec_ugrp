from django import http
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.base import ContentFile
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
import base64
from typing import Any, Dict
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, DetailView
from django.contrib.auth.decorators import login_required
from django.http.response import FileResponse, HttpResponse
from results.models import (Semester, SemesterEnroll, Department, ExamCommittee,
                            Session, Course, Backup, StudentCustomDocument, StudentAcademicData)
from account.models import StudentAccount, AdminAccount
from results.pdf_generators.gradesheet_generator import get_gradesheet
from results.pdf_generators.transcript_generator import render_transcript_for_student
from results.utils import (get_ordinal_number, render_error, get_ordinal_suffix, get_pk_from_base64,
                           has_semester_access, has_semester_super_access)
from results.pdf_generators.course_report_generator import render_coursereport
from results.pdf_generators.coursemedium_v2 import render_coursemedium_cert
from results.pdf_generators.appeared_cert_v2 import render_appeared_cert
from results.pdf_generators.testimonial_generator import render_testimonial
from results.pdf_generators.scorelist_generator import render_scorelist
from results.pdf_generators.topsheet_generator import render_topsheet
from results.pdf_generators.customdoc_generator import render_customdoc_document
from results.pdf_generators.utils import merge_pdfs_from_buffers
from results.decorators_and_mixins import (admin_required, 
                                           superadmin_required, 
                                           superadmin_or_deptadmin_required,
                                           SuperOrDeptAdminRequiredMixin,
                                           DeptAdminRequiredMixin,
                                           SuperAdminRequiredMixin, SuperAdminOrDeptHeadsRequiredMixin)
from .data_processors import get_semester_table_data, get_courseresults_data
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO
from django.conf import settings
from django.core.files.storage import default_storage


def user_is_super_OR_dept_admin(request):
    if hasattr(request.user, 'adminaccount'):
        return request.user.adminaccount.is_super_admin or (request.user.adminaccount.dept is not None)
    else:
        return False

def user_is_super_or_sust_admin(request):
    if hasattr(request.user, 'adminaccount'):
        is_super_admin = request.user.adminaccount.is_super_admin
        is_sust_admin = (request.user.adminaccount.type == 'sust')
        return  (is_super_admin or is_sust_admin)
    else:
        return False

def user_is_superAdmin(user):
    return hasattr(user, 'adminaccount') and user.adminaccount.is_super_admin



@login_required
def homepage(request):
    if not hasattr(request.user, 'adminaccount'):
        return render_error(request, "Unauthorized")
    if request.user.adminaccount.type == "sust":
        return redirect('results:sustadminhome')
    if request.user.adminaccount.type == "academic":
        return redirect('results:secacademichome')
    else:
        # SuperAdmin or dept admin
        context = {}
        context['request'] = request
        department_semesters = []
        departments = Department.objects.all()
        for dept in departments:
            semesters = Semester.objects.filter(session__dept=dept, is_running=True).order_by('session__from_year', "-added_in")
            affiliated_semester_list = []
            for sem in semesters:
                if has_semester_access(sem, request.user.adminaccount):
                    affiliated_semester_list.append(sem)
            department_semesters.append(
                {
                    'name': dept.name.upper(),
                    'semesters': affiliated_semester_list
                }
            )
        context['departments'] = department_semesters
        context['this_is_dept_admin'] = not request.user.adminaccount.is_super_admin
        return render(request, "results/dashboard.html", context=context)
    
    
class SustAdminHome(LoginRequiredMixin, TemplateView):
    template_name = 'sustadmin/build/index.html'   
    def dispatch(self, request, *args: Any, **kwargs: Any) -> HttpResponse:
        if hasattr(request.user, 'adminaccount'):
            if request.user.adminaccount.type == 'sust':
                return super().dispatch(request, *args, **kwargs)
            else:
                return render_error(request, "Forbidden: Only SUST staffs are allowed.")
        else:
            return HttpResponseForbidden("Forbidden: You must be an SUST staff to access this page.")
     
class SecAcademicHome(LoginRequiredMixin, TemplateView):
    template_name = 'sec_academic/build/index.html'
    def dispatch(self, request, *args: Any, **kwargs: Any) -> HttpResponse:
        if hasattr(request.user, 'adminaccount'):
            if request.user.adminaccount.type == 'academic':
                return super().dispatch(request, *args, **kwargs)
            else:
                return render_error(request, "Forbidden: Only academic staffs are allowed.")
        else:
            return HttpResponseForbidden("Forbidden: You must be an academic staff to access this page.")

class DepartmentView(SuperOrDeptAdminRequiredMixin, DetailView):
    template_name = "results/view_department.html"
    
    def get_dept(self):
        return self.get_object()
    
    def get_object(self):
        department = get_object_or_404(Department, name=self.kwargs.get("dept_name", ""))
        return department
    
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context =  super().get_context_data(**kwargs)
        context['request'] = self.request
        context['session_add_access'] = self.request.user.adminaccount.is_super_admin or (self.request.user.adminaccount == context['department'].head)
        return context
    
class ExtensionsView(LoginRequiredMixin, TemplateView):
    template_name = "results/extensions.html"   
     
class GradesheetMakerView(LoginRequiredMixin, TemplateView):
    template_name = "results/gradesheetmaker.html"     

class TranscriptMakerView(LoginRequiredMixin, TemplateView):
    template_name = "results/transcriptmaker.html"  

class CustomdocMakerView(SuperAdminOrDeptHeadsRequiredMixin, TemplateView):
    template_name = "results/customdocmaker.html"
    

@superadmin_or_deptadmin_required
def departments_all(request):
    context = {
            "departments": Department.objects.all(),
            "request": request
        }
    return render(request, "results/departments_all.html", context=context)


class SessionView(SuperOrDeptAdminRequiredMixin, DetailView):
    template_name = "results/view_session.html"
    
    def get_dept(self):
        return self.get_object().dept
    
    def get_object(self):
        session = get_object_or_404(
            Session, 
            dept__name = self.kwargs.get("dept_name", ""),
            from_year = self.kwargs.get("from_year", ""),
            to_year = self.kwargs.get("to_year", ""),
        )
        return session
    
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context =  super().get_context_data(**kwargs)
        context['request'] = self.request
        session = self.get_object()
        semesters = session.semester_set.all().order_by('semester_no', 'repeat_number', 'part_no')
        semesters_context = []
        admin_ac = self.request.user.adminaccount
        for sem in semesters:
            has_acceess = False
            if has_semester_access(sem, admin_ac):
                has_acceess = True
            semesters_context.append(
                {
                    'semester': sem,
                    'has_access': has_acceess
                }
            )
        context['edit_access'] = admin_ac.is_super_admin or (session.dept.head == admin_ac)
        context['semesters'] = semesters_context
        return context


class SemesterView(SuperOrDeptAdminRequiredMixin, DetailView):
    template_name = "results/view_semester.html"
    
    def get_dept(self):
        return self.get_object().session.dept
    
    def get_object(self):
        pk = get_pk_from_base64(self.kwargs.get("b64_id", ""))
        if pk is None:
            raise Http404()
        semester = get_object_or_404(
            Semester, 
            session__dept__name = self.kwargs.get("dept_name", ""),
            id = pk,
        )
        if not has_semester_access(semester, self.request.user.adminaccount):
            raise PermissionDenied
        return semester
    
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context =  super().get_context_data(**kwargs)
        semester = context['semester']
        committee = self.request.COOKIES.get("committee", "hidden")
        if committee == "shown":
            context['is_committee_shown'] = True
        context['request'] = self.request
        context['courses_regular'] = semester.course_set.filter(is_carry_course=False).order_by('id')
        context['courses_created_drop'] = semester.course_set.filter(is_carry_course=True).order_by('id')
        context['courses_drop'] = semester.drop_courses.all().order_by('id')
        context['editor_access'] = has_semester_access(semester, self.request.user.adminaccount)
        context['super_access'] = has_semester_super_access(semester, self.request.user.adminaccount)
        # drop courses semester for current semester
        
        if semester.is_running:
            drop_semesters = Semester.objects.filter(is_running = True, 
                                                     year__lt = semester.year, 
                                                     session__from_year__gt = semester.session.from_year, 
                                                     session__dept = semester.session.dept).order_by("session__from_year")
            if len(drop_semesters) > 0:
                context["drop_semesters"] = drop_semesters
            
        return context
    

class CourseView(DeptAdminRequiredMixin, DetailView):
    template_name = "results/view_course.html"
    
    def get_dept(self):
        return self.get_object().semester.session.dept
    
    def get_object(self):
        pk = get_pk_from_base64(self.kwargs.get("b64_id", ""))
        if pk is None:
            raise Http404()
        course = get_object_or_404(
            Course, 
            semester__session__dept__name = self.kwargs.get("dept_name", ""),
            id = pk
        )
        from_semester_pk = self.request.GET.get('sem')
        if from_semester_pk is not None:
            from_semester = get_object_or_404(Semester, pk=from_semester_pk)
            if not has_semester_access(from_semester, self.request.user.adminaccount):
                raise PermissionDenied
        elif not has_semester_access(course.semester, self.request.user.adminaccount):
            raise PermissionDenied
        return course
    
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context =  super().get_context_data(**kwargs)
        context['request'] = self.request
        course = context['course']
        context['from_session'] = self.request.GET.get('from')
        from_semester_pk = self.request.GET.get('sem')
        if from_semester_pk is not None:
            context['from_semester'] = get_object_or_404(Semester, pk=from_semester_pk)
        dept_running_semesters = Semester.objects.filter(
            is_running = True,
            session__dept = course.semester.session.dept,
        ).order_by('session__from_year')
        context['running_semesters'] = dept_running_semesters.exclude(pk=course.semester.id)
        admin_ac = self.request.user.adminaccount
        context['editor_access'] = admin_ac in course.semester.editor_members
        return context
    

class StaffsView(SuperOrDeptAdminRequiredMixin, TemplateView):
    template_name = "results/view_staffs.html"
    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['superadmins'] = AdminAccount.objects.filter(is_super_admin=True, user__is_staff=False) 
        context['deptadmins'] = AdminAccount.objects.filter(dept__isnull=False).order_by('dept') 
        context['sustdmins'] = AdminAccount.objects.filter(type='sust') 
        context['secacademics'] = AdminAccount.objects.filter(type='academic') 
        context['sysadmins'] = AdminAccount.objects.filter(user__is_staff=True) 
        context['request'] = self.request
        context['departments'] = Department.objects.all()
        return context
    

@admin_required
def download_semester_tabulation(request, pk):
    semester = get_object_or_404(Semester, pk=pk)
    if semester.has_tabulation_sheet:
        filepath = semester.semesterdocument.tabulation_sheet.path
        filename = semester.semesterdocument.tabulation_filename
        return FileResponse(open(filepath, 'rb'), filename=filename)
    else:
        return render_error(request, 'No Tabulation sheet')


@admin_required
def download_year_gradesheet(request, registration, year):
    has_permission = user_is_super_or_sust_admin(request)
    if not has_permission:
        return render_error(request, 'Forbidden')
    student = get_object_or_404(StudentAccount, registration=registration)
    # semester enrolls
    try:
        year_first_semester = SemesterEnroll.objects.get(student=student, semester__year=year, semester__year_semester=1, semester__is_running=False, semester_gpa__isnull=False)
        year_second_semester = SemesterEnroll.objects.get(student=student, semester__year=year, semester__year_semester=2, semester__is_running=False, semester_gpa__isnull=False)
    except:
        return render_error(request, 'GradeSheet not available!')
    sheet_pdf = get_gradesheet(
        student = student,
        year_first_sem_enroll = year_first_semester,
        year_second_sem_enroll = year_second_semester
    )
    filename = f"{get_ordinal_number(year)} year gradesheet - {student.registration}.pdf"
    return FileResponse(ContentFile(sheet_pdf), filename=filename)


@admin_required
def download_semester_gradesheet(request, registration, semester_no):
    has_permission = user_is_super_or_sust_admin(request)
    if not has_permission:
        return render_error(request, 'Forbidden')
    student = get_object_or_404(StudentAccount, registration=registration)
    # semester enrolls
    try:
        semester_enroll = SemesterEnroll.objects.get(student=student, 
                                                        semester__semester_no=semester_no, 
                                                        semester__is_running=False, semester_gpa__isnull=False)
    except:
        return render_error(request, 'GradeSheet not available!')
    sheet_pdf = get_gradesheet(
        student = student,
        year_first_sem_enroll = semester_enroll,
    )
    filename = f"{get_ordinal_number(semester_no)} semester gradesheet - {student.registration}.pdf"
    return FileResponse(ContentFile(sheet_pdf), filename=filename)

@admin_required
def download_transcript(request, registration):
    has_permission = user_is_super_or_sust_admin(request)
    if not has_permission:
        return render_error(request, 'Forbidden')
    try:
        sheet_pdf = render_transcript_for_student(request, registration)
    except Exception as e:
        return render_error(request, e)
    filename = f"Transcript - {registration}.pdf"
    return FileResponse(ContentFile(sheet_pdf), filename=filename)
    


@admin_required
def download_full_document(request, registration):
    has_permission = user_is_super_or_sust_admin(request)
    if not has_permission:
        return render_error(request, 'Forbidden')
    student = get_object_or_404(StudentAccount, registration=registration)
    docs = []
    custom_docs = StudentCustomDocument.objects.filter(student=student, doc_type='all_gss').first()
    if custom_docs:
        docs.append(render_customdoc_document(custom_docs, request))
    gradesheets_semesters = student.gradesheet_semesters
    gradesheets_semester_groups = [gradesheets_semesters[i:i+2] for i in range(0, len(gradesheets_semesters), 2)]
    for year_semester_list in gradesheets_semester_groups:
        year_first_semester = None
        year_second_semester = None
        try:
            year_first_semester = SemesterEnroll.objects.get(student=student, semester__semester_no=year_semester_list[0])
            if len(year_semester_list) == 2:
                year_second_semester = SemesterEnroll.objects.get(student=student, semester__semester_no=year_semester_list[1])
        except:
            render_error(request, 'All the gradesheets of the document is not available!')
        docs.append(get_gradesheet(
            student = student,
            year_first_sem_enroll = year_first_semester,
            year_second_sem_enroll = year_second_semester
        ))
    repeat_semesterenrolls = student.repeat_semester_enrolls
    for enroll in repeat_semesterenrolls:
        docs.append(get_gradesheet(student, year_first_sem_enroll=enroll))
    merged_pdf = merge_pdfs_from_buffers(docs).getvalue()
    filename = f"Transcript & Gradesheets - {registration}.pdf"
    return FileResponse(ContentFile(merged_pdf), filename=filename)
    


@admin_required
def download_coursemediumcert(request, registration):
    student_acadoc = StudentAcademicData.objects.filter(registration=registration).first()
    extra_info = {
        'ref': request.GET.get('ref'),
        'date_today': timezone.now(),
        'admin_name': request.user.adminaccount.user_full_name
    }
    if student_acadoc:
        context = {**student_acadoc.data, **extra_info}
    elif student:=StudentAccount.objects.filter(pk=registration).first():
        context = {
            'name': student.student_name,
            'registration': student.registration,
            'session': student.session.session_code_formal,
            'dept': student.session.dept.fullname,
            'cgpa': student.student_cgpa,
            **extra_info
        }
    sheet_pdf = render_coursemedium_cert(context)
    filename = f"CourseMedium Certificate - {registration}.pdf"
    return FileResponse(ContentFile(sheet_pdf), filename=filename)
    

@admin_required
def download_appeared_cert(request, registration):
    student_acadoc = StudentAcademicData.objects.filter(registration=registration).first()
    duration_str = ''
    if student_acadoc:
        context = student_acadoc.data
        context['semester_no'] = context['semesters_completed']
        context['semester_suffix'] = get_ordinal_suffix(context['semester_no'])
        context['completed_years'] = context.get('years_completed', '')
        if duration := context.get('exam_duration'):
            duration_str = duration
    elif student:=StudentAccount.objects.filter(pk=registration).first():
        last_enroll = student.semesterenroll_set.all().order_by('-semester__semester_no').first()
        if last_enroll is not None:
            last_sesmester_number = last_enroll.semester.semester_no
            context = {
                'name': student.student_name,
                'father_name': student.father_name,
                'mother_name': student.mother_name,
                'registration': student.registration,
                'session': student.session.session_code_formal,
                'dept': student.session.dept.fullname,
                'completed_years': last_sesmester_number//2,
                'semester_no': last_sesmester_number,
                'semester_suffix': get_ordinal_suffix(last_sesmester_number),
            }
            if hasattr(last_enroll.semester, 'duration_info'):
                duration_str = last_enroll.semester.duration_info
        else:
            return render_error(request, 'Appearance Certificate not available without a semester participated!')    
    else:
        return render_error(request, 'Appearance Certificate not available')
    context['admin_name'] = request.user.adminaccount.user_full_name
    context['date_today'] = timezone.now()
    context['ref'] = request.GET.get('ref')
    context['publication_date'] = request.GET.get('pub_date', 'April 30, 2024')
    duration_from = duration_str.split('to')[0].strip()
    duration_to = duration_str.split('to')[-1].strip()
    context['duration_from'] = duration_from
    context['duration_to'] = duration_to
    sheet_pdf = render_appeared_cert(context)
    filename = f"Appeared Certificate - {registration}.pdf"
    return FileResponse(ContentFile(sheet_pdf), filename=filename)
    


@admin_required
def download_testimonial(request, registration):
    student_acadoc = StudentAcademicData.objects.filter(registration=registration).first()
    extra_info = {
        'ref': request.GET.get('ref'),
        'date_today': timezone.now(),
        'admin_name': request.user.adminaccount.user_full_name
    }
    if student_acadoc:
        context = {**student_acadoc.data, **extra_info}
    elif student:=StudentAccount.objects.filter(pk=registration).first():
        last_enroll = student.semesterenroll_set.all().order_by('-semester__semester_no').first()
        if last_enroll is not None:
            last_sem = last_enroll.semester
            last_sem_no = last_sem.semester_no
            context = {
                'name': student.student_name,
                'registration': student.registration,
                'father_name': student.father_name,
                'mother_name': student.mother_name,
                'session': student.session.session_code_formal,
                'dept': student.session.dept.fullname,
                'years_completed': last_sem_no//2,
                'semesters_completed': last_sem_no,
                'exam': last_sem.semester_name,
                'exam_held_in': last_sem.held_in,
                'cgpa': student.student_cgpa,
                **extra_info
            }
        else:
            return render_error(request, 'Testimonial not available without a semester participated!')
    else:
        return render_error(request, 'Testimonial not available')
    sheet_pdf = render_testimonial(context)
    filename = f"Testimonial - {registration}.pdf"
    return FileResponse(ContentFile(sheet_pdf), filename=filename)


@superadmin_required
def download_backup(request, pk):
    has_permission = user_is_super_OR_dept_admin(request)
    if not has_permission:
        return render_error(request, 'Forbidden')
    backup = get_object_or_404(Backup, pk=pk) 
    if request.user.adminaccount.dept is not None and backup.department != request.user.adminaccount.dept:
        return render_error(request, 'Forbidden')
    creation_datetime = backup.created_at
    formatted_datetime = creation_datetime.strftime("%c").replace(" ", "_").replace(":", "-")
    if backup.session:
        filename = f"RDB Backup-{backup.session.batch_name} {formatted_datetime}"
    else:
        filename = f"RDB Backup-{backup.department.name.upper()} {formatted_datetime}"
    response = JsonResponse(backup.data)
    response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
    return response


@superadmin_or_deptadmin_required
def download_coruse_report(request, b64_id):
    try:
        str_pk = base64.b64decode(b64_id.encode('utf-8')).decode()
        pk = int(str(str_pk))
    except Exception as e:
        return render_error(request, "Invalid Course ID")
    course = get_object_or_404(Course, pk=pk)
    from_session_pk = str(request.GET.get('from')).strip()
    from_session = None
    if from_session_pk.isdigit():
        from_session = get_object_or_404(Session, pk=from_session_pk)
    report_pdf = render_coursereport(course, from_session)
    filename = f"{str(course)} Report.pdf"
    return FileResponse(ContentFile(report_pdf), filename=filename)


@superadmin_or_deptadmin_required
def download_scorelist(request, b64_id):
    examiner_name = request.GET.get('examiner', '')
    examiner_designation = request.GET.get('designation', '')
    if any([len(examiner_name) == 0, len(examiner_designation) == 0]):
        return render_error(request, "Examiner Info Missing")
    try:
        str_pk = base64.b64decode(b64_id.encode('utf-8')).decode()
        pk = int(str(str_pk))
    except Exception as e:
        return render_error(request, "Invalid Course ID")
    course = get_object_or_404(Course, pk=pk)
    # report_pdf = render_scorelist(course, examiner_name, examiner_designation)
    report_pdf = render_topsheet(course)
    filename = f"Sust-ScoreList {str(course)}.pdf"
    return FileResponse(ContentFile(report_pdf), filename=filename)
    
    
@login_required
def download_cachedpdf(request, cache_key, filename):
    pdf_base64 = cache.get(cache_key)
    if pdf_base64:
        pdf_data = base64.b64decode(pdf_base64.encode('utf-8'))
        return FileResponse(ContentFile(pdf_data), filename=filename)
    else:
        return render_error(request, "File not found in Cache")
        
    
    
@login_required 
def pending_view(request):
    return render_error(request, 'This Section is Under Development!')

@admin_required
def download_customdoc_template(request):
    file_name = "customdoc_template.xlsx"
    file_path = settings.BASE_DIR / ('results/template_files/'+file_name)
    return FileResponse(
        open(file_path, 'rb'), 
        content_type='application/vnd.ms-excel', 
        filename=file_name, as_attachment=True
    )
    
@admin_required
def download_academicdoc_template(request):
    file_name = "academic_manual_data_template.xlsx"
    file_path = settings.BASE_DIR / ('results/template_files/'+file_name)
    return FileResponse(
        open(file_path, 'rb'), 
        content_type='application/vnd.ms-excel', 
        filename=file_name, as_attachment=True
    )
    
@admin_required
def download_sustdocs_template(request, pk):
    file_name = "supplementary_docs_template.xlsx"
    file_path = settings.BASE_DIR / ('results/template_files/'+file_name)
    course = get_object_or_404(Course, pk=pk)
    file_name = course.code.replace(' ', '') + "_" + file_name
    return FileResponse(
        open(file_path, 'rb'), 
        content_type='application/vnd.ms-excel', 
        filename=file_name, as_attachment=True
    )

@superadmin_or_deptadmin_required
def get_course_excel(request, b64_id):
    try:
        str_pk = base64.b64decode(b64_id.encode('utf-8')).decode()
        pk = int(str(str_pk))
    except Exception as e:
        return render_error(request, "Invalid Course ID")
    course = get_object_or_404(Course, pk=pk)
    from_session_pk = str(request.GET.get('from')).strip()
    from_session = None
    if from_session_pk.isdigit():
        from_session = get_object_or_404(Session, pk=from_session_pk)
    data = get_courseresults_data(course, from_session)
    workbook = Workbook()
    worksheet = workbook.active
    for row_index, row_data in enumerate(data):
        for column_index, cell_value in enumerate(row_data):
            worksheet.cell(row=row_index + 1, column=column_index + 1, value=cell_value)
    # Stylings
    num_cols = worksheet.max_column
    for i in range(num_cols):
        worksheet.column_dimensions[chr(ord('A')+i)].width = 15
        if i > 25:
            break
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    buffer = BytesIO()
    workbook.save(buffer)
    filename = f'{str(course)}.xlsx'
    return FileResponse(
        ContentFile(buffer.getvalue()),
        content_type='application/vnd.ms-excel', 
        filename=filename, as_attachment=True
    )
    

@superadmin_or_deptadmin_required 
def get_semester_excel(request, pk):
    sem = get_object_or_404(Semester, pk=pk)
    data = get_semester_table_data(sem)
    workbook = Workbook()
    worksheet = workbook.active
    for row_index, row_data in enumerate(data):
        for column_index, cell_value in enumerate(row_data):
            worksheet.cell(row=row_index + 1, column=column_index + 1, value=cell_value)
    # Stylings
    num_cols = worksheet.max_column
    for i in range(num_cols):
        if i == 0:
            worksheet.column_dimensions[chr(ord('A')+i)].width = 10
        elif i == 1:
            worksheet.column_dimensions[chr(ord('A')+i)].width = 35
        else:
            worksheet.column_dimensions[chr(ord('A')+i)].width = 15
        if i > 25:
            break
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    buffer = BytesIO()
    workbook.save(buffer)
    filename = f'semester_data {sem.semester_code}.xlsx'
    return FileResponse(
        ContentFile(buffer.getvalue()),
        content_type='application/vnd.ms-excel', 
        filename=filename, as_attachment=True
    )


@admin_required
def download_customdoc(request, reg, doctype):
    num = request.GET.get('num')
    if num:
        cdoc = StudentCustomDocument.objects.filter(student__registration=reg, doc_type=doctype, sem_or_year_num=num).first()
    else:
        cdoc = StudentCustomDocument.objects.filter(student__registration=reg, doc_type=doctype).first()
    if not cdoc:
        return render_error(request, "No Document Found")
    if not cdoc.document_data:
        return render_error(request, "Document data is empty")
    doc_pdf = render_customdoc_document(cdoc, request)
    return FileResponse(ContentFile(doc_pdf), filename=cdoc.filename)

@admin_required
def download_supplementarydoc(request, b64_id):
    try:
        str_pk = base64.b64decode(b64_id.encode('utf-8')).decode()
        pk = int(str(str_pk))
    except Exception as e:
        return render_error(request, "Invalid Course ID")
    course = get_object_or_404(Course, pk=pk)
    if not hasattr(course, 'supplementarydocument'):
        return render_error(request, 'Document Doesn\'t exists')
    s_doc = course.supplementarydocument
    filepath = s_doc.document.path
    filename = s_doc.document_filename
    return FileResponse(open(filepath, 'rb'), filename=filename)
    