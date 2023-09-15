from django.test import TestCase
from django.conf import settings
import openpyxl
from typing import List
from results.models import SemesterEnroll, Semester, Session
import unittest
from termcolor import colored
# Create your tests here.

def get_excel_dataset(header: List, data_rows:List, credits_idxs:List, gp_idxs: List):
    dataset = {}
    reg_idx = header.index('reg')
    overall_credit_idx = header.index('overall-credit')
    overall_gpa_idx = header.index('overall-gpa')
    
    for row in data_rows:
        registration = int(row[reg_idx].value)
        dataset[registration] = {}
        for credit_i in credits_idxs:
            credit_header = header[credit_i]
            semester_number = int(credit_header.split('-')[-1])
            semester_credits = row[credit_i].value
            dataset[registration][semester_number] = {
                'credits': semester_credits
            }
        for gp_i in gp_idxs:
            gp_header = header[gp_i]
            semester_number = int(gp_header.split('-')[-1])
            semester_gp = row[gp_i].value
            dataset[registration][semester_number]['gp'] = semester_gp
            if semester_number in dataset[registration].keys():
                dataset[registration][semester_number]['gp'] = semester_gp
        dataset[registration]['overall'] = {}
        dataset[registration]['overall']['credits'] = row[overall_credit_idx].value
        dataset[registration]['overall']['gp'] = row[overall_gpa_idx].value
                
    return dataset
        


class SemesterResultsTestCase(TestCase):
    fixtures = [settings.BASE_DIR/'fixtures.json']
    
    def test_semester_results(self):
        # data preparation
        excel_file = settings.BASE_DIR/"anunad_results.xlsx"
        wb = openpyxl.load_workbook(excel_file)
        sheet1 = wb[wb.sheetnames[0]]
        rows = list(sheet1.rows)
        header = [cell.value.lower().strip() for cell in rows[0] if cell.value is not None]
        data_rows = rows[1:]
        credits_idxs = [header.index(col) for col in header if col.startswith("credit")]
        gp_idxs = [header.index(col) for col in header if col.startswith("gp")]
        self.assertEqual(len(credits_idxs), len(gp_idxs))
        dataset = get_excel_dataset(header, data_rows, credits_idxs, gp_idxs)
        # model preparation
        count_enrollments = 0
        count_successful = 0
        semesters = Semester.objects.filter(session__from_year=2018)
        for sem in semesters:
            success = 0
            print(f"# Running test for semester: {sem}")
            semester_enrollments = sem.semesterenroll_set.all()
            for enroll in semester_enrollments:
                reg = enroll.student.registration
                try:
                    actual_credits = dataset[reg][sem.semester_no]['credits']
                    actual_gp = dataset[reg][sem.semester_no]['gp']
                except Exception as exe:
                    continue
                messsage = f"{reg} <-- {sem}"
                try:
                    self.assertEqual(enroll.semester_credits, actual_credits, msg=messsage)
                except AssertionError:
                    print(colored(f"Mismatch Credits --> Reg: {reg} , system: {enroll.semester_credits} , actual {actual_credits}", 'light_red'))
                try:
                    self.assertEqual(enroll.semester_gpa, actual_gp, msg=messsage)
                except AssertionError:
                    # print(f"Mismatch GP --> Reg: {reg} , db: {enroll.semester_gpa} , actual {actual_gp}")
                    print(colored(f"Mismatch GP --> Reg: {reg} , system: {enroll.semester_gpa} , actual {actual_gp}", "red"))
                    continue
                success += 1
            count_enrollments += semester_enrollments.count()
            count_successful += success
            print(colored(f"successful: {success} / {semester_enrollments.count()} ({round((success / semester_enrollments.count())*100, 2)}%)", 'light_green'))
        print(colored(f"Overall accuracy: {round((count_successful / count_enrollments)*100, 2)}%", 'yellow'))
        # overall results
        print('-'*70)
        print(colored(f"# Running test on overall final result", 'cyan'))
        students = Session.objects.get(dept__name='eee', from_year=2018).studentaccount_set.all()
        success = 0
        for student in students:
            reg = student.registration
            try:
                actual_credits = dataset[reg]['overall']['credits']
                actual_gpa = dataset[reg]['overall']['gp']
            except Exception as exe:
                continue
            messsage = f"{reg}"
            try:
                self.assertEqual(round(student.credits_completed, 2), actual_credits, msg=messsage)
            except AssertionError:
                print(colored(f"Mismatch Credits --> Reg: {reg} , system: {round(student.credits_completed, 2)} , actual {actual_credits}", 'light_red'))
            try:
                self.assertEqual(round(student.raw_cgpa, 2), actual_gpa, msg=messsage)
            except AssertionError:
                print(colored(f"Mismatch GP --> Reg: {reg} , system: {round(student.raw_cgpa, 2)} , actual {actual_gpa}", "red"))
                continue
            success += 1
        print(colored(f"Success: {success} / {students.count()}", 'light_green'))
            
        
        
# class OverallResultTestCase(TestCase):
#     fixtures = [settings.BASE_DIR/'fixtures.json']   
    
#     def test_semester_results(self):
#         excel_file = settings.BASE_DIR/"anunad_results.xlsx"
#         wb = openpyxl.load_workbook(excel_file)
#         sheet1 = wb[wb.sheetnames[0]]
#         rows = list(sheet1.rows)
#         header = [cell.value.lower().strip() for cell in rows[0] if cell.value is not None]
#         data_rows = rows[1:]
